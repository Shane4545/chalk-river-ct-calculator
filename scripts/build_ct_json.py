"""Build ct-tables.json from Chalk River CT Calculator.xlsx Giardia layout.

Each row's k[0] is the 0.5 C column (matches Calculations!N8 VLOOKUP to Giardia CT $S$4:$X$18).
Other k[] indices are extra temperature columns from the sheet; the web calculator uses k[0] only.
"""
import json
import openpyxl

path = r"c:\Cursor Projects\CT Calculator\Chalk River CT Calculator.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Giardia CT"]

# Blocks: row 4 + 16*n for n=0..4 -> pH on row+1 col I
blocks = []
for n in range(30):
    base = 4 + 16 * n
    ph = ws.cell(base + 1, 9).value
    if ph is None:
        break
    if not isinstance(ph, (int, float)):
        continue
    temps_row = [ws.cell(base, 9 + 2 * i).value for i in range(5)]
    if temps_row[0] != 0.5:
        continue
    pairs = []
    for dr in range(2, 16):
        r = base + dr
        row_cl = ws.cell(r, 10).value
        if row_cl is None or not isinstance(row_cl, (int, float)):
            break
        ks = []
        for ti in range(5):
            k = ws.cell(r, 9 + 2 * ti).value
            ks.append(float(k) if k is not None else None)
        pairs.append({"cl": float(row_cl), "k": ks})
    blocks.append({"ph": float(ph), "temps": [0.5, 5, 10, 15, 20], "rows": pairs})

out = {
    "giardia_blocks": blocks,
    "virus_k": [
        [0.5, 0.3333333333333333],
        [5, 0.5],
        [10, 0.6944444444444443],
        [15, 1.0],
        [20, 1.611111111111111],
        [25, 2.3333333333333335],
    ],
    "segment_names": ["Clearwell", "Pipe CW to Tower", "Elevated tower"],
}

out_path = r"c:\Cursor Projects\CT Calculator\ct-tables.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(out, f, indent=2)

js_path = r"c:\Cursor Projects\CT Calculator\ct-tables.js"
with open(js_path, "w", encoding="utf-8") as f:
    f.write("window.CT_TABLES = ")
    json.dump(out, f)
    f.write(";\n")

print("Wrote", out_path, "and", js_path, "blocks", len(blocks))
for b in blocks:
    print(" pH", b["ph"], "rows", len(b["rows"]))

wb.close()
