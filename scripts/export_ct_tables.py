"""Export Giardia k tables per temperature from workbook (cached values)."""
import json
import openpyxl

path = r"c:\Cursor Projects\CT Calculator\Chalk River CT Calculator.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Giardia CT"]

# pH row is 3: cols T through X in S-block = 7,7.5,8,8.5,9 (S is Co, T-X pH)
ph_levels = [7, 7.5, 8, 8.5, 9]
chlorine_levels = []
# S5:S18 chlorine
for r in range(5, 19):
    chlorine_levels.append(ws.cell(r, 19).value)

# HLOOKUP blocks: each temp has 16 rows; temps in row 4 of each block col I,K,M...
# Read I column for block starts
blocks = []
for base in range(4, 2000, 16):
    temp_cell = ws.cell(base, 9).value  # col I
    if temp_cell is None and base > 100:
        break
    # chlorine in col I rows base+1 to base+14, k in J through N for 5 pH? 
    # Check: row 5 had S=0.4 - that's col 19. So block might be I-Q not S-X
    row0 = [ws.cell(base + i, 9).value for i in range(1, 16)]
    if row0[0] is None:
        continue
    # sample row base+1 across 9-17
    sample = [ws.cell(base + 1, c).value for c in range(9, 18)]
    blocks.append((base, temp_cell, sample))

# Simpler: use columns S-X only — values already temperature-specific in saved file
# Re-load with different approach: parse row 1 "Temperature 1:" and sibling cells

t1_row = 1
# Actually multiple temperature tables exist as HLOOKUP sources in I:Q
# Export I4:Q83 as flat structure
data = {}
row = 4
while row < 500:
    t = ws.cell(row, 9).value
    if t is None and row > 200:
        break
    # If this row looks like a header (0.5 in col I and number in K)
    if isinstance(t, (int, float)) and t in (0.5, 5, 10, 15, 20, 25):
        temp_key = float(t)
        matrix = []
        for dr in range(1, 16):
            r = row + dr
            cl = ws.cell(r, 9).value
            if cl is None:
                break
            ks = [ws.cell(r, 10 + j).value for j in range(0, 9, 2)]  # J,L,N,P,R? 
            # pH 7,7.5,8,8.5,9 might be cols J,N,R not every other
        row += 1
        continue
    row += 1

wb.close()

# Minimal export: S5:X18 k table + virus k (already have)
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["Giardia CT"]
giardia_sx = []
for r in range(5, 19):
    giardia_sx.append([ws.cell(r, c).value for c in range(19, 25)])

out = {
    "giardia_k_sx": giardia_sx,
    "ph_cols": ph_levels,
    "virus_k": [
        [0.5, 0.3333333333333333],
        [5, 0.5],
        [10, 0.6944444444444443],
        [15, 1],
        [20, 1.611111111111111],
        [25, 2.3333333333333335],
    ],
}
print(json.dumps(out, indent=2)[:2000])
